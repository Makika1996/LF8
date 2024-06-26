import tkinter as tk
from tkinter import filedialog, messagebox, scrolledtext
import mariadb
import csv
import xml.etree.ElementTree as ET
from openpyxl import Workbook
from openpyxl.utils import get_column_letter

# Verbindungsparameter zur Datenbank definieren
config = {
    'user': 'root',
    #'password': '123',
    #'host': '10.145.240.100',
    'host':'127.0.0.1',
    'port': 3306,
    #'database': 'Heiner'
    'database':'heiner_it'
}

# Funktion zur Auswahl der CSV-Datei und zum Schreiben der Daten
def select_and_write_to_csv(query):
    try:
        # Verbindung zur Datenbank herstellen
        conn = mariadb.connect(**config)
        cursor = conn.cursor()

        # SQL-Abfrage ausführen
        cursor.execute(query)

        # Alle Zeilen der Abfrage abrufen
        rows = cursor.fetchall()

        # Datei auswählen und speichern
        filename = filedialog.asksaveasfilename(defaultextension=".csv", filetypes=[("CSV files", "*.csv")])
        if filename:
            # CSV-Datei öffnen und Daten schreiben
            with open(filename, 'w', newline='', encoding='utf-8') as file:
                csv_writer = csv.writer(file)
                
                # Spaltenüberschriften schreiben
                csv_writer.writerow([i[0] for i in cursor.description])
                
                # Daten aus der Datenbank in die CSV-Datei schreiben
                csv_writer.writerows(rows)

            messagebox.showinfo("Erfolg", f"Daten erfolgreich in {filename} geschrieben.")
        else:
            messagebox.showwarning("Warnung", "Keine Datei ausgewählt.")

    except mariadb.Error as e:
        messagebox.showerror("Fehler", f"Fehler bei der Datenbankabfrage: {e}")

    finally:
        # Verbindung schließen
        if conn:
            conn.close()

# Funktion zum Exportieren der Ergebnisse als XML
def write_to_xml(query):
    try:
        # Verbindung zur Datenbank herstellen
        conn = mariadb.connect(**config)
        cursor = conn.cursor()

        # SQL-Abfrage ausführen
        cursor.execute(query)

        # Alle Zeilen der Abfrage abrufen
        rows = cursor.fetchall()

        # XML-Element erstellen
        root = ET.Element("results")

        # Ergebnisse als XML-Elemente hinzufügen
        for row in rows:
            result = ET.SubElement(root, "result")
            for idx, col in enumerate(cursor.description):
                field = ET.SubElement(result, col[0])
                field.text = str(row[idx])

        # XML-Datei speichern
        filename = filedialog.asksaveasfilename(defaultextension=".xml", filetypes=[("XML files", "*.xml")])
        if filename:
            tree = ET.ElementTree(root)
            tree.write(filename, encoding='utf-8', xml_declaration=True)
            messagebox.showinfo("Erfolg", f"Daten erfolgreich in {filename} als XML gespeichert.")
        else:
            messagebox.showwarning("Warnung", "Keine Datei ausgewählt.")

    except mariadb.Error as e:
        messagebox.showerror("Fehler", f"Fehler bei der Datenbankabfrage: {e}")

    finally:
        # Verbindung schließen
        if conn:
            conn.close()

# Funktion zum Exportieren der Ergebnisse als Excel
def write_to_excel(query):
    try:
        # Verbindung zur Datenbank herstellen
        conn = mariadb.connect(**config)
        cursor = conn.cursor()

        # SQL-Abfrage ausführen
        cursor.execute(query)

        # Alle Zeilen der Abfrage abrufen
        rows = cursor.fetchall()

        # Excel-Datei erstellen
        wb = Workbook()
        ws = wb.active
        ws.title = "Query Results"

        # Spaltenüberschriften schreiben
        for col_num, col in enumerate(cursor.description, 1):
            ws.cell(row=1, column=col_num, value=col[0])

        # Daten aus der Datenbank in die Excel-Datei schreiben
        for row_num, row in enumerate(rows, 2):
            for col_num, value in enumerate(row, 1):
                ws.cell(row=row_num, column=col_num, value=value)

        # Excel-Datei speichern
        filename = filedialog.asksaveasfilename(defaultextension=".xlsx", filetypes=[("Excel files", "*.xlsx")])
        if filename:
            wb.save(filename)
            messagebox.showinfo("Erfolg", f"Daten erfolgreich in {filename} als Excel gespeichert.")
        else:
            messagebox.showwarning("Warnung", "Keine Datei ausgewählt.")

    except mariadb.Error as e:
        messagebox.showerror("Fehler", f"Fehler bei der Datenbankabfrage: {e}")

    finally:
        # Verbindung schließen
        if conn:
            conn.close()

# Funktion zur Anzeige der GUI für die Abfrage und Auswahl der Exportoptionen
def show_gui_for_query_and_export():
    # Hauptfenster erstellen
    root = tk.Tk()
    root.title("Datenbankabfrage und Export")
    root.geometry("600x400")
    root.configure(bg="light grey")

    # SQL-Abfrage Textfeld
    query_label = tk.Label(root, text="SQL-Abfrage:", font=("Helvetica", 14), bg="light grey")
    query_label.pack()

    query_entry = scrolledtext.ScrolledText(root, width=70, height=5, wrap=tk.WORD, bg="light blue")
    query_entry.pack(pady=10)

    # Funktion zum Ausführen der SQL-Abfrage und Exportoptionen anzeigen
    def execute_query_and_show_export_options():
        query = query_entry.get("1.0", tk.END).strip()
        if not query:
            messagebox.showwarning("Warnung", "Bitte geben Sie eine SQL-Abfrage ein.")
            return
        
        # Buttons für Export hinzufügen
        export_csv_button = tk.Button(root, text="Export als CSV", command=lambda: select_and_write_to_csv(query), bg="light blue")
        export_csv_button.pack(pady=5)

        export_xml_button = tk.Button(root, text="Export als XML", command=lambda: write_to_xml(query), bg="light blue")
        export_xml_button.pack(pady=5)

        export_excel_button = tk.Button(root, text="Export als Excel", command=lambda: write_to_excel(query), bg="light blue")
        export_excel_button.pack(pady=5)

    # Binde die Enter-Taste an die Funktion execute_query_and_show_export_options
    query_entry.bind("<Return>", lambda event: execute_query_and_show_export_options())

    # Ausführen-Button
    execute_button = tk.Button(root, text="Abfrage ausführen", command=execute_query_and_show_export_options, bg="light blue")
    execute_button.pack(pady=5)

    root.mainloop()

# Funktion aufrufen, um die GUI anzuzeigen und die Funktionen bereitzustellen
show_gui_for_query_and_export()
